#!/usr/bin/env python3
import sys
import json
import math
import openpyxl
from pathlib import Path

# xlwings はエコリング用に使用（書式・画像を完全に維持）
try:
    import xlwings as xw
    XLWINGS_AVAILABLE = True
except ImportError:
    XLWINGS_AVAILABLE = False

def generate_starbuyers_bag(items, template_path, output_path):
    """スターバイヤーズ バッグ出品リストを生成"""
    wb = openpyxl.load_workbook(template_path)
    ws = wb['出品リスト (バッグ)']

    # 既存データをクリア（13行目以降、余裕を持って200行まで）
    for row in range(13, max(ws.max_row + 1, 200)):
        for col in range(1, 9):
            ws.cell(row=row, column=col, value=None)

    # 新しいデータを書き込み
    for i, item in enumerate(items):
        row = 13 + i
        # 指値 = 仕入総額 + 1万円（千円単位で切り上げ）
        sashi_ne = item.get('purchase_total')
        if sashi_ne:
            sashi_ne = math.ceil((sashi_ne + 10000) / 1000) * 1000

        # 商品名 = ブランド名 + 商品名
        brand_name = item.get('brand_name', '')
        product_name = item.get('product_name', '')
        if brand_name and product_name:
            full_product_name = f"{brand_name}　{product_name}"
        elif brand_name:
            full_product_name = brand_name
        else:
            full_product_name = product_name

        ws.cell(row=row, column=1, value=i + 1)  # No
        ws.cell(row=row, column=2, value=full_product_name)  # 商品名
        ws.cell(row=row, column=3, value=item.get('condition_rank', 'B'))  # ランク
        ws.cell(row=row, column=4, value=item.get('accessories', ''))  # 付属品
        ws.cell(row=row, column=5, value='')  # 備考（空）
        ws.cell(row=row, column=6, value=sashi_ne)  # 指値
        ws.cell(row=row, column=7, value='')  # ロット番号
        ws.cell(row=row, column=8, value=item.get('management_number', ''))  # 管理番号

    wb.save(output_path)
    return True

def generate_starbuyers_accessory(items, template_path, output_path):
    """スターバイヤーズ アクセサリー出品リストを生成"""
    wb = openpyxl.load_workbook(template_path)

    # シートを探す
    sheet_name = None
    for name in wb.sheetnames:
        if '出品リスト' in name:
            sheet_name = name
            break

    if not sheet_name:
        sheet_name = wb.sheetnames[-1]

    ws = wb[sheet_name]

    # 既存データをクリア（13行目以降、余裕を持って200行まで）
    for row in range(13, max(ws.max_row + 1, 200)):
        for col in range(1, 9):
            ws.cell(row=row, column=col, value=None)

    # 新しいデータを書き込み
    for i, item in enumerate(items):
        row = 13 + i
        # 指値 = 仕入総額 + 1万円（千円単位で切り上げ）
        sashi_ne = item.get('purchase_total')
        if sashi_ne:
            sashi_ne = math.ceil((sashi_ne + 10000) / 1000) * 1000

        # 商品名 = ブランド名 + 商品名
        brand_name = item.get('brand_name', '')
        product_name = item.get('product_name', '')
        if brand_name and product_name:
            full_product_name = f"{brand_name}　{product_name}"
        elif brand_name:
            full_product_name = brand_name
        else:
            full_product_name = product_name

        ws.cell(row=row, column=1, value=i + 1)
        ws.cell(row=row, column=2, value=full_product_name)
        ws.cell(row=row, column=3, value=item.get('condition_rank', 'B'))
        ws.cell(row=row, column=4, value=item.get('accessories', ''))
        ws.cell(row=row, column=5, value='')  # 備考（空）
        ws.cell(row=row, column=6, value=sashi_ne)
        ws.cell(row=row, column=7, value='')
        ws.cell(row=row, column=8, value=item.get('management_number', ''))

    wb.save(output_path)
    return True

def generate_ecoring_brand(items, template_path, output_path):
    """エコリング ブランド出品リストを生成（xlwings使用で書式完全維持）"""
    if not XLWINGS_AVAILABLE:
        raise Exception("xlwings is not installed. Please run: pip install xlwings")

    # Excelアプリを非表示で起動
    app = xw.App(visible=False)
    try:
        wb = app.books.open(template_path)
        ws = wb.sheets['委託者入力シート']

        # 新しいデータを書き込み（16行目から）
        for i, item in enumerate(items):
            row = 16 + i

            # 指値 = 仕入総額 + 1万円（千円単位で切り上げ）
            sashi_ne = item.get('purchase_total')
            if sashi_ne:
                sashi_ne = math.ceil((sashi_ne + 10000) / 1000) * 1000

            # 商品名 = ブランド名 + 商品名
            brand_name = item.get('brand_name', '')
            product_name = item.get('product_name', '')
            if brand_name and product_name:
                full_product_name = f"{brand_name} {product_name}"
            elif brand_name:
                full_product_name = brand_name
            else:
                full_product_name = product_name

            ws.range(f'A{row}').value = i + 1  # 商品NO
            ws.range(f'B{row}').value = full_product_name  # 商品名
            ws.range(f'C{row}').value = sashi_ne  # 指値
            ws.range(f'D{row}').value = ''  # ダメージ・備考（空）
            ws.range(f'E{row}').value = item.get('management_number', '')  # メモ欄

        wb.save(output_path)
        wb.close()
    finally:
        app.quit()

    return True

def generate_ecoring_dougu(items, template_path, output_path):
    """エコリング 道具出品リストを生成（xlwings使用で書式完全維持）"""
    if not XLWINGS_AVAILABLE:
        raise Exception("xlwings is not installed. Please run: pip install xlwings")

    # Excelアプリを非表示で起動
    app = xw.App(visible=False)
    try:
        wb = app.books.open(template_path)
        ws = wb.sheets['委託者入力シート']

        # 新しいデータを書き込み（16行目から）
        for i, item in enumerate(items):
            row = 16 + i

            # 指値 = 仕入総額 + 1万円（千円単位で切り上げ）
            sashi_ne = item.get('purchase_total')
            if sashi_ne:
                sashi_ne = math.ceil((sashi_ne + 10000) / 1000) * 1000

            # 商品名 = ブランド名 + 商品名
            brand_name = item.get('brand_name', '')
            product_name = item.get('product_name', '')
            if brand_name and product_name:
                full_product_name = f"{brand_name} {product_name}"
            elif brand_name:
                full_product_name = brand_name
            else:
                full_product_name = product_name

            ws.range(f'A{row}').value = i + 1  # 商品NO
            ws.range(f'B{row}').value = full_product_name  # 商品名
            ws.range(f'C{row}').value = sashi_ne  # 指値
            ws.range(f'D{row}').value = ''  # ダメージ・備考（空）
            ws.range(f'E{row}').value = item.get('management_number', '')  # メモ欄

        wb.save(output_path)
        wb.close()
    finally:
        app.quit()

    return True

def generate_appre_brand(items, template_path, output_path):
    """アプレオークション ブランド出品リストを生成（xlwings使用で書式・マクロ完全維持）"""
    if not XLWINGS_AVAILABLE:
        raise Exception("xlwings is not installed. Please run: pip install xlwings")

    # Excelアプリを非表示で起動
    app = xw.App(visible=False)
    try:
        wb = app.books.open(template_path)
        ws = wb.sheets['入力欄']

        # 新しいデータを書き込み（6行目から）
        for i, item in enumerate(items):
            row = 6 + i

            # 指値 = 仕入総額 + 1万円（千円単位で切り上げ）
            sashi_ne = item.get('purchase_total')
            if sashi_ne:
                sashi_ne = math.ceil((sashi_ne + 10000) / 1000) * 1000

            # ブランド名と商品名は別々の列
            brand_name = item.get('brand_name', '')
            product_name = item.get('product_name', '')

            ws.range(f'E{row}').value = brand_name  # ブランド名（列5）
            ws.range(f'G{row}').value = product_name  # 商品名（列7）
            ws.range(f'K{row}').value = item.get('condition_rank', 'B')  # ランク（列11）
            ws.range(f'L{row}').value = ''  # 付属品・備考（空）
            ws.range(f'N{row}').value = sashi_ne  # 指値（税抜）（列14）
            ws.range(f'T{row}').value = item.get('management_number', '')  # 管理番号（列20）

        wb.save(output_path)
        wb.close()
    finally:
        app.quit()

    return True

def main():
    if len(sys.argv) < 2:
        print(json.dumps({'error': 'No input provided'}))
        sys.exit(1)

    try:
        input_data = json.loads(sys.argv[1])
        items = input_data.get('items', [])
        auction_type = input_data.get('auctionType', 'starbuyers-bag')
        template_dir = input_data.get('templateDir', '')
        output_path = input_data.get('outputPath', '')

        if auction_type == 'starbuyers-bag':
            template_path = Path(template_dir) / 'starbuyers_bag_template.xlsx'
            generate_starbuyers_bag(items, str(template_path), output_path)
        elif auction_type == 'starbuyers-accessory':
            template_path = Path(template_dir) / 'starbuyers_accessory_template.xlsx'
            generate_starbuyers_accessory(items, str(template_path), output_path)
        elif auction_type == 'ecoring-brand':
            template_path = Path(template_dir) / 'ecoring_brand_template.xlsx'
            generate_ecoring_brand(items, str(template_path), output_path)
        elif auction_type == 'ecoring-dougu':
            template_path = Path(template_dir) / 'ecoring_dougu_template.xlsx'
            generate_ecoring_dougu(items, str(template_path), output_path)
        elif auction_type == 'appre-brand':
            template_path = Path(template_dir) / 'appre_brand_template.xlsm'
            generate_appre_brand(items, str(template_path), output_path)
        else:
            print(json.dumps({'error': f'Unknown auction type: {auction_type}'}))
            sys.exit(1)

        print(json.dumps({'success': True, 'outputPath': output_path}))

    except Exception as e:
        print(json.dumps({'error': str(e)}))
        sys.exit(1)

if __name__ == '__main__':
    main()
